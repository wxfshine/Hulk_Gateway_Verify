import requests
import logging
import sys
import atexit
import os
import csv
import html
import io
import json
import webbrowser
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import argparse
from datetime import datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from urllib.parse import unquote

from api_case_testing import run_api_case_testing

# 1. 基础配置
base_url = "http://hulk.cmit.local:18080/v1"  # 你的LLM Gateway地址
api_key = "eyJhbGciOiJIUzUxMiIsImlhdCI6MTc3OTA5MDczOSwiZXhwIjoxNzg2ODY2NzM5fQ.eyJpZCI6NzE2LCJuYW1lIjoid2FuZ3hmIiwic291cmNlIjoibGRhcCIsImV4dGVybmFsX2lkIjoie2VjYWE5ZGVkLWI2MzMtNGNmOC05Y2Q5LWQxZGE4MTYyYWFhYn0iLCJyb2xlIjoiVXNlciJ9.M_lPymDGllUI0BaMULupNQ8HaWa9G0zYfRgm5AGz8KBaBuBTcru472vOZilgULdROivMsRJmCxVlNkL81nvlfw"  # 从POST /v1/api-keys拿到的Key

EMBEDDING_MODEL_NAMES = {
    "dengcao/Qwen3-Embedding-4B:Q4_K_M",
    "text-embedding-nomic-embed-text-v1.5"
}

CHAT_MODEL_NAMES = {
    "Qwen3.6-35B-A3B",
    "deepseek-ocr",
    "gemma-4-31B-it",
    "gemma4-26b-a4b",
    "llama3.2",
    "qwen3-14b-awq"
}

RESULT_HEADERS = [
    "timestamp",
    "status",
    "model_name",
    "compliance_check",
    "input_content",
    "api_name",
    # 额外字段，方便调试：预期/实际状态码、请求体、响应头、响应体
    "expected_http_status",
    "actual_http_status",
    "request_payload",
    "response_headers",
    "response_body",
    "result_or_error"
]


# 默认 contents 配置（当 docs 中没有配置文件时使用）
DEFAULT_CONTENTS = {
    "chat_completion": [
        "介绍一下Python",
        "你能做些什么",
        "我今天非常开心",
        "中国历史上最强大的王朝是哪个？",
        "美国哪一年成立?",
        "请给我说说人工智能中大模型、边缘模型是什么意思"
    ],
    "completion": [
        "春天来了，万物复苏",
        "今天天气真好，适合出去游玩",
        "人工智能的发展趋势是",
        "历史上最惨烈的水库泄洪事故",
        "北京通州未来房价趋势",
        "21世纪人类航天可能会有哪些突破"
    ],
    "embedding": [
        "我喜欢AI",
        "今天要下雨吗",
        "航天科技的发展趋势是"
    ]
}


def load_contents_config(file_path: str = None):
    """从指定的 JSON 配置文件中加载 contents 配置。
    文件示例 (docs/test_contents.json):
    {
      "chat_completion": ["...", "..."],
      "completion": ["..."],
      "embedding": ["..."]
    }
    """
    if not file_path:
        file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "docs", "test_contents.json")

    try:
        if os.path.isfile(file_path):
            with open(file_path, "r", encoding="utf-8") as fh:
                data = json.load(fh)
                # 保证所有三类存在
                return {
                    "chat_completion": data.get("chat_completion", DEFAULT_CONTENTS["chat_completion"]),
                    "completion": data.get("completion", DEFAULT_CONTENTS["completion"]),
                    "embedding": data.get("embedding", DEFAULT_CONTENTS["embedding"])
                }
        else:
            print(f"未找到 contents 配置文件: {file_path}，使用内置默认配置。")
            return DEFAULT_CONTENTS.copy()
    except Exception as err:
        print(f"读取 contents 配置失败: {err}，使用内置默认配置。")
        return DEFAULT_CONTENTS.copy()


class TeeStream:
    def __init__(self, *streams):
        self.streams = streams

    def write(self, message):
        for stream in self.streams:
            stream.write(message)
            stream.flush()

    def flush(self):
        for stream in self.streams:
            stream.flush()


def setup_logging():
    log_dir = get_log_dir()
    os.makedirs(log_dir, exist_ok=True)
    log_file = os.path.join(log_dir, f"hulk_gateway_verify_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log")
    log_handle = open(log_file, "w", encoding="utf-8")

    original_stdout = sys.stdout
    original_stderr = sys.stderr

    sys.stdout = TeeStream(original_stdout, log_handle)
    sys.stderr = TeeStream(original_stderr, log_handle)

    def cleanup_logging():
        sys.stdout = original_stdout
        sys.stderr = original_stderr
        log_handle.close()

    atexit.register(cleanup_logging)

    def handle_exception(exc_type, exc_value, exc_traceback):
        if issubclass(exc_type, KeyboardInterrupt):
            original_stderr.write("程序被手动中断。\n")
            original_stderr.flush()
            log_handle.write("程序被手动中断。\n")
            log_handle.flush()
            return

        logging.error("程序发生未捕获异常", exc_info=(exc_type, exc_value, exc_traceback))

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s - %(levelname)s - %(message)s",
        handlers=[logging.StreamHandler(sys.stdout)]
    )
    sys.excepthook = handle_exception

    print(f"日志文件: {log_file}")
    return log_handle


def create_result_writer():
    log_dir = get_log_dir()
    os.makedirs(log_dir, exist_ok=True)
    result_file = os.path.join(log_dir, f"hulk_gateway_verify_result_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv")
    result_handle = open(result_file, "w", newline="", encoding="utf-8-sig")
    writer = csv.writer(result_handle)
    writer.writerow(RESULT_HEADERS)
    result_handle.flush()

    def cleanup_result_file():
        result_handle.close()

    atexit.register(cleanup_result_file)
    print(f"结果表格文件: {result_file}")
    return writer, result_handle, result_file


def format_result(result):
    if isinstance(result, list):
        preview = result[:10]
        return f"向量长度: {len(result)}, 前10项: {preview}"
    return str(result)


def record_test_result(
    writer,
    result_handle,
    status,
    model_name,
    compliance_check,
    input_content,
    api_name,
    result_or_error,
    expected_http_status="",
    actual_http_status="",
    request_payload="",
    response_headers="",
    response_body=""
):
    """将一条测试结果写入 CSV。额外的网络元数据可选传入。"""
    writer.writerow([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        status,
        model_name,
        compliance_check,
        input_content,
        api_name,
        expected_http_status,
        actual_http_status,
        request_payload,
        response_headers,
        response_body,
        result_or_error
    ])
    result_handle.flush()


def get_log_dir():
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), "log")


def read_result_rows(reader, source_name):
    if not reader.fieldnames:
        raise ValueError(f"结果文件缺少表头: {source_name}")

    missing_headers = [header for header in RESULT_HEADERS if header not in reader.fieldnames]
    if missing_headers:
        raise ValueError(f"结果文件字段缺失 {missing_headers}: {source_name}")

    rows = []
    for row in reader:
        if not any((value or "").strip() for value in row.values() if isinstance(value, str)):
            continue
        rows.append({header: row.get(header, "") for header in RESULT_HEADERS})
    return rows


def create_merged_result_file(rows, source_descriptions=None):
    log_dir = get_log_dir()
    os.makedirs(log_dir, exist_ok=True)
    merged_file = os.path.join(log_dir, f"hulk_gateway_verify_merged_result_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv")

    sorted_rows = sorted(rows, key=lambda row: (row.get("timestamp", ""), row.get("model_name", ""), row.get("api_name", "")))

    with open(merged_file, "w", newline="", encoding="utf-8-sig") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=RESULT_HEADERS)
        writer.writeheader()
        writer.writerows(sorted_rows)

    if source_descriptions:
        print(f"已合并 {len(source_descriptions)} 个结果文件，输出文件: {merged_file}")

    return merged_file


def merge_result_files(file_paths):
    rows = []
    for file_path in file_paths:
        with open(file_path, "r", encoding="utf-8-sig", newline="") as csv_file:
            reader = csv.DictReader(csv_file)
            rows.extend(read_result_rows(reader, file_path))

    if not rows:
        raise ValueError("未从指定结果文件中读取到有效测试数据。")

    return create_merged_result_file(rows, source_descriptions=file_paths)


def export_csv_to_excel_with_failure_details(csv_path):
    """读取 CSV（script 输出），生成一个带“失败用例明细”表格的 XLSX 文件。返回 xlsx 路径。"""
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Summary"

    # 读取 CSV 行
    with open(csv_path, "r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh)
        rows = list(reader)

    # 写入 Summary（这里简单写入总数/成功/失败）
    total = len(rows)
    success = sum(1 for r in rows if r.get("status") == "success")
    error = total - success
    summary_ws.append(["报告生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    summary_ws.append(["总测试数", total])
    summary_ws.append(["成功数", success])
    summary_ws.append(["失败数", error])

    # 失败用例明细表
    detail_ws = wb.create_sheet(title="失败用例明细")
    # 五列：预期状态码/实际状态码/请求体/响应头/响应体，并保留基本信息
    headers = [
        "timestamp",
        "model_name",
        "api_name",
        "input_content",
        "expected_http_status",
        "actual_http_status",
        "request_payload",
        "response_headers",
        "response_body",
        "result_or_error"
    ]
    detail_ws.append(headers)

    for r in rows:
        if r.get("status") != "success":
            detail_ws.append([
                r.get("timestamp", ""),
                r.get("model_name", ""),
                r.get("api_name", ""),
                r.get("input_content", ""),
                r.get("expected_http_status", ""),
                r.get("actual_http_status", ""),
                r.get("request_payload", ""),
                r.get("response_headers", ""),
                r.get("response_body", ""),
                r.get("result_or_error", "")
            ])

    # 格式化列宽和样式
    for ws in (summary_ws, detail_ws):
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 80)
            ws.column_dimensions[col_letter].width = adjusted_width

    # 保存 xlsx
    xlsx_path = os.path.splitext(csv_path)[0] + ".xlsx"
    wb.save(xlsx_path)
    return xlsx_path


def merge_result_contents(file_items):
    rows = []
    source_names = []

    for item in file_items:
        source_name = item.get("name") or "未命名结果文件"
        content = item.get("content", "")
        if not isinstance(content, str) or not content.strip():
            raise ValueError(f"结果文件内容为空: {source_name}")

        reader = csv.DictReader(io.StringIO(content))
        rows.extend(read_result_rows(reader, source_name))
        source_names.append(source_name)

    if not rows:
        raise ValueError("未从上传结果文件中读取到有效测试数据。")

    return create_merged_result_file(rows, source_descriptions=source_names)


def analyze_test_results(result_file):
    analysis = {
        "report_generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "test_start_time": "",
        "test_end_time": "",
        "total_count": 0,
        "success_count": 0,
        "error_count": 0,
        "overall_success_rate": 0.0,
        "api_stats": {},
        "model_stats": {},
        "compliance_stats": {},
        "error_type_stats": {},
        "input_error_stats": {},
        "feedback_stats": {
            "total_count": 0,
            "success_count": 0,
            "error_count": 0,
            "success_rate": 0.0
        },
        "mismatch_records": [],
        "summary": []
    }

    with open(result_file, "r", encoding="utf-8-sig", newline="") as csv_file:
        reader = csv.DictReader(csv_file)
        rows = list(reader)

    if rows:
        analysis["test_start_time"] = rows[0]["timestamp"]
        analysis["test_end_time"] = rows[-1]["timestamp"]

    for row in rows:
        status = row["status"]
        model_name = row["model_name"]
        api_name = row["api_name"]
        compliance_check = str(row["compliance_check"])
        input_content = row["input_content"]
        result_or_error = row["result_or_error"]

        analysis["total_count"] += 1
        if status == "success":
            analysis["success_count"] += 1
        else:
            analysis["error_count"] += 1

        api_stats = analysis["api_stats"].setdefault(api_name, {"total": 0, "success": 0, "error": 0, "success_rate": 0.0})
        api_stats["total"] += 1
        api_stats[status] += 1

        model_stats = analysis["model_stats"].setdefault(model_name, {"total": 0, "success": 0, "error": 0, "success_rate": 0.0})
        model_stats["total"] += 1
        model_stats[status] += 1

        compliance_stats = analysis["compliance_stats"].setdefault(compliance_check, {"total": 0, "success": 0, "error": 0, "success_rate": 0.0})
        compliance_stats["total"] += 1
        compliance_stats[status] += 1

        is_embedding_api = api_name == "验证调用向量"
        is_embedding_model = model_name in EMBEDDING_MODEL_NAMES
        is_chat_model = model_name in CHAT_MODEL_NAMES
        if (is_embedding_api and not is_embedding_model) or ((not is_embedding_api) and not is_chat_model):
            analysis["mismatch_records"].append({
                "model_name": model_name,
                "api_name": api_name,
                "input_content": input_content,
                "status": status
            })

        if api_name == "验证调用反馈提交与查询":
            analysis["feedback_stats"]["total_count"] += 1
            if status == "success":
                analysis["feedback_stats"]["success_count"] += 1
            else:
                analysis["feedback_stats"]["error_count"] += 1

        if status == "error":
            error_type = result_or_error.split(":", 1)[0] if ":" in result_or_error else "UnknownError"
            analysis["error_type_stats"][error_type] = analysis["error_type_stats"].get(error_type, 0) + 1
            analysis["input_error_stats"][input_content] = analysis["input_error_stats"].get(input_content, 0) + 1

    if analysis["total_count"]:
        analysis["overall_success_rate"] = analysis["success_count"] / analysis["total_count"] * 100

    for stats in analysis["api_stats"].values():
        if stats["total"]:
            stats["success_rate"] = stats["success"] / stats["total"] * 100

    for stats in analysis["model_stats"].values():
        if stats["total"]:
            stats["success_rate"] = stats["success"] / stats["total"] * 100

    for stats in analysis["compliance_stats"].values():
        if stats["total"]:
            stats["success_rate"] = stats["success"] / stats["total"] * 100

    if analysis["feedback_stats"]["total_count"]:
        analysis["feedback_stats"]["success_rate"] = analysis["feedback_stats"]["success_count"] / analysis["feedback_stats"]["total_count"] * 100

    best_model = None
    worst_model = None
    if analysis["model_stats"]:
        sorted_models = sorted(
            analysis["model_stats"].items(),
            key=lambda item: (-item[1]["success_rate"], -item[1]["success"], item[0])
        )
        best_model = sorted_models[0]
        worst_model = sorted(
            analysis["model_stats"].items(),
            key=lambda item: (item[1]["success_rate"], -item[1]["error"], item[0])
        )[0]

    analysis["summary"].append(
        f"本次共执行 {analysis['total_count']} 条测试，成功 {analysis['success_count']} 条，失败 {analysis['error_count']} 条，整体成功率 {analysis['overall_success_rate']:.2f}% 。"
    )

    if best_model:
        analysis["summary"].append(
            f"成功率最高的模型是 {best_model[0]}，成功率 {best_model[1]['success_rate']:.2f}% 。"
        )

    if worst_model:
        analysis["summary"].append(
            f"成功率最低的模型是 {worst_model[0]}，成功率 {worst_model[1]['success_rate']:.2f}% 。"
        )

    if analysis["feedback_stats"]["total_count"]:
        analysis["summary"].append(
            f"反馈链路共测试 {analysis['feedback_stats']['total_count']} 条，成功率 {analysis['feedback_stats']['success_rate']:.2f}% 。"
        )

    if analysis["error_type_stats"]:
        top_error = sorted(analysis["error_type_stats"].items(), key=lambda item: (-item[1], item[0]))[0]
        analysis["summary"].append(
            f"最常见错误类型为 {top_error[0]}，共出现 {top_error[1]} 次。"
        )

    if analysis["mismatch_records"]:
        analysis["summary"].append(
            f"发现 {len(analysis['mismatch_records'])} 条模型类型与接口类型不匹配的记录，需要进一步检查测试分流逻辑。"
        )
    else:
        analysis["summary"].append("未发现模型类型与接口类型不匹配的记录。")

    return analysis


def generate_analysis_html(analysis, result_file, report_title="Hulk Gateway 测试结果分析报告"):
    report_file = os.path.join(
        os.path.dirname(result_file),
        f"hulk_gateway_verify_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
    )

    def build_stats_rows(stats_dict, key_title):
        if not stats_dict:
            return "<tr><td colspan='5'>暂无数据</td></tr>"
        rows = []
        for name, stats in sorted(stats_dict.items(), key=lambda item: item[0]):
            rows.append(
                f"<tr><td>{html.escape(str(name))}</td><td>{stats['total']}</td><td>{stats['success']}</td><td>{stats['error']}</td><td>{stats['success_rate']:.2f}%</td></tr>"
            )
        return "".join(rows)

    def build_simple_rows(stats_dict, value_title):
        if not stats_dict:
            return f"<tr><td colspan='2'>暂无{html.escape(value_title)}</td></tr>"
        rows = []
        for name, count in sorted(stats_dict.items(), key=lambda item: (-item[1], item[0])):
            rows.append(
                f"<tr><td>{html.escape(str(name))}</td><td>{count}</td></tr>"
            )
        return "".join(rows)

    def build_mismatch_rows(records):
        if not records:
            return "<tr><td colspan='4'>未发现不匹配记录</td></tr>"
        rows = []
        for record in records:
            rows.append(
                f"<tr><td>{html.escape(record['model_name'])}</td><td>{html.escape(record['api_name'])}</td><td>{html.escape(record['input_content'])}</td><td>{html.escape(record['status'])}</td></tr>"
            )
        return "".join(rows)

    summary_html = "".join(f"<li>{html.escape(item)}</li>" for item in analysis["summary"])

    html_content = f"""
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <title>{html.escape(report_title)}</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 24px; background: #f7f9fc; color: #1f2937; }}
        h1, h2 {{ color: #111827; }}
        .card {{ background: #ffffff; border-radius: 8px; padding: 20px; margin-bottom: 20px; box-shadow: 0 2px 8px rgba(0, 0, 0, 0.08); }}
        table {{ width: 100%; border-collapse: collapse; margin-top: 12px; }}
        th, td {{ border: 1px solid #d1d5db; padding: 8px 10px; text-align: left; vertical-align: top; }}
        th {{ background: #eef2ff; }}
        ul {{ margin: 0; padding-left: 20px; }}
        .metric {{ display: inline-block; min-width: 180px; margin-right: 16px; margin-bottom: 8px; }}
        .path {{ color: #2563eb; word-break: break-all; }}
    </style>
</head>
<body>
    <div class="card">
        <h1>{html.escape(report_title)}</h1>
        <p class="path">结果文件：{html.escape(result_file)}</p>
        <div class="metric">报告生成时间：{html.escape(analysis['report_generated_at'])}</div>
        <div class="metric">测试开始时间：{html.escape(analysis['test_start_time'] or '暂无数据')}</div>
        <div class="metric">测试结束时间：{html.escape(analysis['test_end_time'] or '暂无数据')}</div>
        <br>
        <div class="metric">总测试数：{analysis['total_count']}</div>
        <div class="metric">成功数：{analysis['success_count']}</div>
        <div class="metric">失败数：{analysis['error_count']}</div>
        <div class="metric">整体成功率：{analysis['overall_success_rate']:.2f}%</div>
    </div>

    <div class="card">
        <h2>测试结论摘要</h2>
        <ul>{summary_html}</ul>
    </div>

    <div class="card">
        <h2>按接口类型统计</h2>
        <table>
            <tr><th>接口类型</th><th>总数</th><th>成功</th><th>失败</th><th>成功率</th></tr>
            {build_stats_rows(analysis['api_stats'], '接口类型')}
        </table>
    </div>

    <div class="card">
        <h2>按模型统计</h2>
        <table>
            <tr><th>模型名称</th><th>总数</th><th>成功</th><th>失败</th><th>成功率</th></tr>
            {build_stats_rows(analysis['model_stats'], '模型名称')}
        </table>
    </div>

    <div class="card">
        <h2>按合规开关统计</h2>
        <table>
            <tr><th>compliance_check</th><th>总数</th><th>成功</th><th>失败</th><th>成功率</th></tr>
            {build_stats_rows(analysis['compliance_stats'], 'compliance_check')}
        </table>
    </div>

    <div class="card">
        <h2>反馈接口专项统计</h2>
        <div class="metric">反馈测试数：{analysis['feedback_stats']['total_count']}</div>
        <div class="metric">反馈成功数：{analysis['feedback_stats']['success_count']}</div>
        <div class="metric">反馈失败数：{analysis['feedback_stats']['error_count']}</div>
        <div class="metric">反馈成功率：{analysis['feedback_stats']['success_rate']:.2f}%</div>
    </div>

    <div class="card">
        <h2>错误类型分布</h2>
        <table>
            <tr><th>错误类型</th><th>出现次数</th></tr>
            {build_simple_rows(analysis['error_type_stats'], '错误类型')}
        </table>
    </div>

    <div class="card">
        <h2>高频失败输入</h2>
        <table>
            <tr><th>输入内容</th><th>失败次数</th></tr>
            {build_simple_rows(analysis['input_error_stats'], '失败输入')}
        </table>
    </div>

    <div class="card">
        <h2>模型类型与接口类型匹配检查</h2>
        <table>
            <tr><th>模型名称</th><th>接口类型</th><th>输入内容</th><th>状态</th></tr>
            {build_mismatch_rows(analysis['mismatch_records'])}
        </table>
    </div>
</body>
</html>
"""

    with open(report_file, "w", encoding="utf-8") as html_file:
        html_file.write(html_content)

    return report_file


def open_analysis_html(report_file):
    try:
        opened = webbrowser.open(f"file://{os.path.abspath(report_file)}")
        if opened:
            print(f"已在默认浏览器中打开分析报告: {report_file}")
        else:
            print(f"未能自动打开浏览器，请手动打开分析报告: {report_file}")
    except Exception as err:
        print(f"自动打开分析报告失败: {type(err).__name__}: {err}")


def build_merge_analysis_portal_html():
    return """<!DOCTYPE html>
<html lang=\"zh-CN\">
<head>
    <meta charset=\"UTF-8\">
    <title>Hulk Gateway 多结果文件合并分析</title>
    <style>
        body { font-family: Arial, sans-serif; margin: 24px; background: #f7f9fc; color: #1f2937; }
        .card { background: #ffffff; border-radius: 8px; padding: 20px; margin-bottom: 20px; box-shadow: 0 2px 8px rgba(0, 0, 0, 0.08); }
        h1, h2 { color: #111827; }
        button { background: #2563eb; color: #fff; border: none; border-radius: 6px; padding: 10px 18px; cursor: pointer; }
        button:disabled { background: #93c5fd; cursor: not-allowed; }
        #status { white-space: pre-wrap; line-height: 1.6; }
        ul { padding-left: 20px; }
        .tip { color: #4b5563; }
        a { color: #2563eb; }
    </style>
</head>
<body>
    <div class=\"card\">
        <h1>Hulk Gateway 多结果文件合并分析</h1>
        <p class=\"tip\">请选择一个或多个测试结果 CSV 文件，点击按钮后会先合并数据，再调用现有分析逻辑生成统一分析报告。</p>
    </div>

    <div class=\"card\">
        <h2>选择结果文件</h2>
        <input id=\"resultFiles\" type=\"file\" accept=\".csv\" multiple>
        <button id=\"analyzeBtn\" type=\"button\">生成数据分析</button>
        <ul id=\"fileList\"></ul>
    </div>

    <div class=\"card\">
        <h2>执行状态</h2>
        <div id=\"status\">等待选择结果文件。</div>
    </div>

    <script>
        const resultFilesInput = document.getElementById('resultFiles');
        const analyzeBtn = document.getElementById('analyzeBtn');
        const fileList = document.getElementById('fileList');
        const status = document.getElementById('status');

        resultFilesInput.addEventListener('change', () => {
            fileList.innerHTML = '';
            const files = Array.from(resultFilesInput.files || []);
            if (!files.length) {
                status.textContent = '等待选择结果文件。';
                return;
            }

            for (const file of files) {
                const li = document.createElement('li');
                li.textContent = `${file.name} (${file.size} bytes)`;
                fileList.appendChild(li);
            }
            status.textContent = `已选择 ${files.length} 个结果文件，等待开始分析。`;
        });

        analyzeBtn.addEventListener('click', async () => {
            const files = Array.from(resultFilesInput.files || []);
            if (!files.length) {
                status.textContent = '请至少选择一个结果文件。';
                return;
            }

            analyzeBtn.disabled = true;
            status.textContent = '正在读取结果文件并提交分析，请稍候...';

            try {
                const filePayload = await Promise.all(
                    files.map(async (file) => ({
                        name: file.name,
                        content: await file.text()
                    }))
                );

                const response = await fetch('/analyze', {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({ files: filePayload })
                });

                const result = await response.json();
                if (!response.ok || !result.success) {
                    throw new Error(result.error || '分析失败');
                }

                status.innerHTML = `分析完成。<br>合并结果文件：${result.merged_file}<br>分析报告：<a href="${result.report_url}" target="_blank">点击打开</a>`;
                window.open(result.report_url, '_blank');
            } catch (error) {
                status.textContent = `分析失败：${error.message}`;
            } finally {
                analyzeBtn.disabled = false;
            }
        });
    </script>
</body>
</html>
"""


class MergeAnalysisRequestHandler(BaseHTTPRequestHandler):
    def do_GET(self):
        if self.path in ["/", "/index.html"]:
            content = build_merge_analysis_portal_html().encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Content-Length", str(len(content)))
            self.end_headers()
            self.wfile.write(content)
            return

        if self.path.startswith("/reports/"):
            report_name = os.path.basename(unquote(self.path[len("/reports/"):]))
            report_file = os.path.join(get_log_dir(), report_name)
            if os.path.isfile(report_file):
                with open(report_file, "rb") as html_file:
                    content = html_file.read()
                self.send_response(200)
                self.send_header("Content-Type", "text/html; charset=utf-8")
                self.send_header("Content-Length", str(len(content)))
                self.end_headers()
                self.wfile.write(content)
                return

        if self.path == "/favicon.ico":
            self.send_response(204)
            self.end_headers()
            return

        self.send_error(404, "未找到请求资源")

    def do_POST(self):
        if self.path != "/analyze":
            self.send_error(404, "未找到请求资源")
            return

        try:
            content_length = int(self.headers.get("Content-Length", "0"))
            request_body = self.rfile.read(content_length)
            payload = json.loads(request_body.decode("utf-8"))
            file_items = payload.get("files", [])
            if not isinstance(file_items, list) or not file_items:
                raise ValueError("请至少上传一个结果文件。")

            merged_file = merge_result_contents(file_items)
            analysis = analyze_test_results(merged_file)
            report_file = generate_analysis_html(analysis, merged_file, report_title="Hulk Gateway 多结果文件合并分析报告")

            response = {
                "success": True,
                "merged_file": merged_file,
                "report_file": report_file,
                "report_url": f"/reports/{os.path.basename(report_file)}"
            }
            response_data = json.dumps(response, ensure_ascii=False).encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Content-Length", str(len(response_data)))
            self.end_headers()
            self.wfile.write(response_data)
        except Exception as err:
            response = {
                "success": False,
                "error": f"{type(err).__name__}: {err}"
            }
            response_data = json.dumps(response, ensure_ascii=False).encode("utf-8")
            self.send_response(400)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Content-Length", str(len(response_data)))
            self.end_headers()
            self.wfile.write(response_data)

    def log_message(self, format, *args):
        logging.info("MergeAnalysisServer - " + format, *args)


def launch_merge_analysis_portal():
    os.makedirs(get_log_dir(), exist_ok=True)
    server = ThreadingHTTPServer(("127.0.0.1", 0), MergeAnalysisRequestHandler)
    host, port = server.server_address
    portal_url = f"http://{host}:{port}/"

    print(f"多结果文件合并分析页面已启动: {portal_url}")
    print("按 Ctrl+C 可停止网页服务。")
    webbrowser.open(portal_url)

    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("多结果文件合并分析页面已停止。")
    finally:
        server.server_close()


def post_json(url, payload):
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json"
    }
    response = requests.post(url, json=payload, headers=headers)
    response.raise_for_status()

    try:
        return response.json()
    except ValueError as err:
        raise ValueError(f"接口返回的不是合法 JSON: {response.text}") from err


def post_json_capture(url, payload):
    """发送请求并返回元数据与解析后的 json（如果可解析）。"""
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json"
    }
    response = requests.post(url, json=payload, headers=headers)
    # 不 raise_for_status，这样可以在非 200 时仍读取响应体
    try:
        parsed = response.json()
    except ValueError:
        parsed = {"raw_text": response.text}

    return {
        "status_code": response.status_code,
        "headers": dict(response.headers),
        "raw_text": response.text,
        "parsed_json": parsed,
        "request_payload": payload
    }


def get_json(url, params=None):
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json"
    }
    response = requests.get(url, headers=headers, params=params)
    response.raise_for_status()

    try:
        return response.json()
    except ValueError as err:
        raise ValueError(f"接口返回的不是合法 JSON: {response.text}") from err


def get_json_capture(url, params=None):
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json"
    }
    response = requests.get(url, headers=headers, params=params)
    try:
        parsed = response.json()
    except ValueError:
        parsed = {"raw_text": response.text}

    return {
        "status_code": response.status_code,
        "headers": dict(response.headers),
        "raw_text": response.text,
        "parsed_json": parsed,
        "request_payload": params
    }


def extract_chat_content(result):
    choices = result.get("choices")
    if not choices:
        raise ValueError(f"接口返回缺少 choices 字段: {result}")

    message = choices[0].get("message")
    if not message or "content" not in message:
        raise ValueError(f"接口返回缺少 message.content 字段: {result}")

    return message["content"]


def extract_completion_text(result):
    choices = result.get("choices")
    if not choices or "text" not in choices[0]:
        raise ValueError(f"接口返回缺少 choices[0].text 字段: {result}")
    return choices[0]["text"]


def extract_embedding_vector(result):
    data = result.get("data")
    if not data or "embedding" not in data[0]:
        raise ValueError(f"接口返回缺少 data[0].embedding 字段: {result}")
    return data[0]["embedding"]


def pretty_headers(headers):
    try:
        return json.dumps(headers, ensure_ascii=False)
    except Exception:
        return str(headers)


def pretty_body(parsed_json, raw_text=None):
    try:
        return json.dumps(parsed_json, ensure_ascii=False)
    except Exception:
        return raw_text or str(parsed_json)


def extract_feedback_items(result):
    if isinstance(result, list):
        return result

    if isinstance(result, dict):
        for key in ["data", "items", "results", "feedbacks"]:
            value = result.get(key)
            if isinstance(value, list):
                return value

    raise ValueError(f"查询反馈接口返回格式不符合预期: {result}")

def get_models(base_url, api_key):
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json"
    }

    response = requests.get(f"{base_url}/models", headers=headers)

    if response.status_code == 200:
        models = response.json()
        print("✅ 获取模型列表成功：")
        for model in models.get("data", []):
            print(f"- 模型ID: {model['id']}")
        return models

    print(f"❌ 请求失败，状态码: {response.status_code}")
    print("错误信息:", response.text)
    return None


def filter_models_by_name(models, allowed_model_names, model_type_name):
    filtered_data = [
        model for model in models.get("data", [])
        if model.get("id") in allowed_model_names
    ]

    print(f"✅ {model_type_name}模型列表：")
    for model in filtered_data:
        print(f"- 模型ID: {model['id']}")

    missing_models = [
        model_name for model_name in allowed_model_names
        if model_name not in {model.get('id') for model in filtered_data}
    ]
    if missing_models:
        print(f"⚠️ 以下{model_type_name}模型未在接口返回列表中找到：")
        for model_name in missing_models:
            print(f"- 模型ID: {model_name}")

    return {"data": filtered_data}



# -------------------- 1. 对话聊天接口（最常用）----------------------
def chat_completion(user_message: str, model_name: str, stream: bool = False,cmp_ck=0, capture: bool = False):
    """
    对话聊天接口
    :param user_message: 用户问题
    :param model_name: 模型名称（必填）
    :param stream: 是否流式输出
    :return: AI 回答内容
    """
    url = f"{base_url}/chat/completions"
    payload = {
        "model": model_name,
        "messages": [{"role": "user", "content": user_message}],
        "stream": stream,
        "compliance_check":cmp_ck
    }

    if capture:
        meta = post_json_capture(url, payload)
        content = extract_chat_content(meta.get("parsed_json", {}))
        return content, meta

    result = post_json(url, payload)
    return extract_chat_content(result)


# -------------------- 2. 文本续写接口 ----------------------
def completion(prompt: str, model_name: str, max_tokens: int = 512,cmp_ck=0, capture: bool = False):
    """
    文本续写接口
    :param prompt: 提示词
    :param model_name: 模型名称（必填）
    :param max_tokens: 最大生成长度
    :return: 续写文本
    """
    url = f"{base_url}/completions"
    payload = {
        "model": model_name,
        "prompt": prompt,
        "max_tokens": max_tokens,
        "compliance_check":cmp_ck
    }

    if capture:
        meta = post_json_capture(url, payload)
        text = extract_completion_text(meta.get("parsed_json", {}))
        return text, meta

    result = post_json(url, payload)
    return extract_completion_text(result)


# -------------------- 3. 向量嵌入接口 ----------------------
def embedding(text: str, model_name: str,cmp_ck=0, capture: bool = False):
    """
    文本转向量（Embedding）
    :param text: 需要转向量的文本
    :param model_name: 向量模型名称（必填）
    :return: 向量数组
    """
    url = f"{base_url}/embeddings"
    payload = {
        "model": model_name,
        "input": text,
        "compliance_check":cmp_ck
    }

    if capture:
        meta = post_json_capture(url, payload)
        vector = extract_embedding_vector(meta.get("parsed_json", {}))
        return vector, meta

    result = post_json(url, payload)
    return extract_embedding_vector(result)


def submit_feedback(messages, model_name: str, rating: int, comment: str = ""):
    url = f"{base_url}/feedback"
    payload = {
        "messages": messages,
        "model": model_name,
        "rating": rating,
        "comment": comment
    }
    return post_json(url, payload)


def query_feedback(page: int = 1, page_size: int = 10):
    url = f"{base_url}/feedback"
    params = {
        "page": page,
        "page_size": page_size
    }
    return get_json(url, params=params)

def run_api_test(models, compliance_check, api_name, contents, api_func, writer, result_handle):
    if not contents:
        print(f"跳过 {api_name}：无待执行的 contents。")
        return

    print(f"-------------开始{api_name}--------------")
    for model in models.get("data", []):
        model_name = model["id"]
        for content in contents:
            print(f"模型名称: {model_name} - 内容输入: {content} - compliance_check: {compliance_check} - 测试项目: {api_name}")
            try:
                # 使用 capture 模式一次请求得到解析值与元数据
                value, meta = api_func(content, model_name=model_name, cmp_ck=compliance_check, capture=True)
                formatted_result = format_result(value)
                print("结果：", formatted_result)
                record_test_result(
                    writer,
                    result_handle,
                    "success",
                    model_name,
                    compliance_check,
                    content,
                    api_name,
                    formatted_result,
                    expected_http_status=200,
                    actual_http_status=meta.get("status_code"),
                    request_payload=meta.get("request_payload"),
                    response_headers=pretty_headers(meta.get("headers")),
                    response_body=pretty_body(meta.get("parsed_json"), meta.get("raw_text"))
                )
            except Exception as err:
                error_message = f"{type(err).__name__}: {err}"
                print("报错信息：", error_message)
                # 如果异常发生，但 meta 可用（部分解析在 api_func 内可能抛出），尽量记录 meta
                try:
                    # 当 api_func 在内部抛出并未返回 meta，我们仍尝试调用 capture separately
                    meta = api_func(content, model_name=model_name, cmp_ck=compliance_check, capture=True)[1]
                except Exception:
                    meta = {}

                record_test_result(
                    writer,
                    result_handle,
                    "error",
                    model_name,
                    compliance_check,
                    content,
                    api_name,
                    error_message,
                    expected_http_status=200,
                    actual_http_status=meta.get("status_code"),
                    request_payload=meta.get("request_payload"),
                    response_headers=pretty_headers(meta.get("headers")) if meta else "",
                    response_body=pretty_body(meta.get("parsed_json"), meta.get("raw_text")) if meta else ""
                )
        print("-" * 60)


def test_feedback(models, compliance_check, writer, result_handle):
    print("-------------开始验证调用反馈提交与查询--------------")
    test_cases = [
        {
            "user_message": "请介绍一下Python的主要特点",
            "rating": 8,
            "comment": f"自动化测试反馈 compliance_check={compliance_check}"
        },
        {
            "user_message": "请总结一下人工智能的发展趋势",
            "rating": 9,
            "comment": f"自动化测试反馈2 compliance_check={compliance_check}"
        }
    ]

    for model in models.get("data", []):
        model_name = model["id"]
        for case in test_cases:
            user_message = case["user_message"]
            rating = case["rating"]
            comment = case["comment"]
            print(f"模型名称: {model_name} - 内容输入: {user_message} - compliance_check: {compliance_check} - 测试项目: 验证调用反馈提交与查询")

            try:
                assistant_reply = chat_completion(user_message, model_name=model_name, cmp_ck=compliance_check)
                messages = [
                    {"role": "user", "content": user_message},
                    {"role": "assistant", "content": assistant_reply}
                ]

                submit_result = submit_feedback(messages, model_name, rating, comment)
                feedback_result = query_feedback(page=1, page_size=10)
                feedback_items = extract_feedback_items(feedback_result)

                matched_feedback = next(
                    (
                        item for item in feedback_items
                        if item.get("model") == model_name
                        and item.get("rating") == rating
                        and item.get("comment") == comment
                        and item.get("messages") == messages
                    ),
                    None
                )

                if not matched_feedback:
                    raise ValueError(
                        f"查询反馈结果中未找到刚提交的数据。submit_result={submit_result}, feedback_result={feedback_result}"
                    )

                formatted_result = f"提交成功并查询校验成功，rating={rating}, comment={comment}"
                print("结果：", formatted_result)
                record_test_result(writer, result_handle, "success", model_name, compliance_check, user_message, "验证调用反馈提交与查询", formatted_result)
            except Exception as err:
                error_message = f"{type(err).__name__}: {err}"
                print("报错信息：", error_message)
                record_test_result(writer, result_handle, "error", model_name, compliance_check, user_message, "验证调用反馈提交与查询", error_message)
        print("-" * 60)


def test_chat_completion(models, compliance_check, writer, result_handle):
    contents = DEFAULT_CONTENTS.get("chat_completion")
    run_api_test(models, compliance_check, "验证调用对话", contents, chat_completion, writer, result_handle)


def test_completion(models, compliance_check, writer, result_handle):
    contents = DEFAULT_CONTENTS.get("completion")
    run_api_test(models, compliance_check, "验证调用续写", contents, completion, writer, result_handle)


def test_embedding(models, compliance_check, writer, result_handle):
    contents = DEFAULT_CONTENTS.get("embedding")
    run_api_test(models, compliance_check, "验证调用向量", contents, embedding, writer, result_handle)


def main():
    setup_logging()
    writer, result_handle, result_file = create_result_writer()
    models = get_models(base_url, api_key)
    print("*" * 20)
    if models:
        chat_models = filter_models_by_name(models, CHAT_MODEL_NAMES, "聊天/对话")
        embedding_models = filter_models_by_name(models, EMBEDDING_MODEL_NAMES, "向量")

        #for compliance_check in [0, 1,2,3]:
        for compliance_check in [2]:
            test_chat_completion(chat_models, compliance_check, writer, result_handle)
            test_completion(chat_models, compliance_check, writer, result_handle)
            test_embedding(embedding_models, compliance_check, writer, result_handle)
            test_feedback(chat_models, compliance_check, writer, result_handle)

        analysis = analyze_test_results(result_file)
        report_file = generate_analysis_html(analysis, result_file)
        print(f"分析报告文件: {report_file}")
        open_analysis_html(report_file)
        print("如需合并多个测试结果文件进行统一分析，可运行: python Hulk_Gateway_Verify.py --merge-report-ui")

        try:
            xlsx_file = export_csv_to_excel_with_failure_details(result_file)
            print(f"已生成 Excel 失败用例明细: {xlsx_file}")
        except Exception as err:
            print(f"生成 Excel 失败用例明细失败: {err}")

    print("✅ 测试流程已执行完毕。")


def parse_args():
    parser = argparse.ArgumentParser(description="Hulk Gateway 接口验证与结果分析工具")
    subparsers = parser.add_subparsers(dest="command")

    api_case_parser = subparsers.add_parser("api-case-run", help="执行 API 用例测试")
    api_case_parser.add_argument("case_file", help="API 用例 CSV 文件路径")

    merge_files_parser = subparsers.add_parser("merge-files", help="直接合并一个或多个测试结果文件并生成分析报告")
    merge_files_parser.add_argument("files", nargs="+", help="测试结果文件路径列表")

    subparsers.add_parser("merge-report-ui", help="启动多测试结果文件合并分析网页")

    parser.add_argument(
        "--merge-report-ui",
        action="store_true",
        help="启动多测试结果文件合并分析网页"
    )
    parser.add_argument(
        "--merge-files",
        nargs="+",
        help="直接合并一个或多个测试结果文件并生成分析报告"
    )
    parser.add_argument(
        "--api-case-file",
        help="执行 API 用例测试，参数为用例 CSV 文件路径"
    )
    parser.add_argument(
        "--contents-file",
        help="指定一个 JSON 文件路径，从中读取 chat_completion/completion/embedding 的 contents 配置（默认: docs/test_contents.json）"
    )
    parser.add_argument(
        "--single-content",
        help="仅运行单个内容项（会在所有对应 API 的模型上执行），输入应当与配置中的某一项文本完全匹配"
    )
    return parser.parse_args()


def run_merge_files_analysis(file_paths):
    merged_file = merge_result_files(file_paths)
    analysis = analyze_test_results(merged_file)
    report_file = generate_analysis_html(analysis, merged_file, report_title="Hulk Gateway 多结果文件合并分析报告")
    print(f"合并结果文件: {merged_file}")
    print(f"分析报告文件: {report_file}")
    open_analysis_html(report_file)

if __name__ == "__main__":
    args = parse_args()
    if args.command == "merge-report-ui" or args.merge_report_ui:
        launch_merge_analysis_portal()
    elif args.command == "merge-files":
        run_merge_files_analysis(args.files)
    elif args.merge_files:
        run_merge_files_analysis(args.merge_files)
    elif args.command == "api-case-run":
        setup_logging()
        setup_logging()
        run_api_case_testing(base_url, api_key, args.case_file, get_log_dir())
        print("✅ API 用例测试流程已执行完毕。")
    elif args.api_case_file:
        setup_logging()
        run_api_case_testing(base_url, api_key, args.api_case_file, get_log_dir())
        print("✅ API 用例测试流程已执行完毕。")
    else:
        # 如果指定了 contents-file，则加载并替换默认配置
        if args.contents_file:
            cfg = load_contents_config(args.contents_file)
            DEFAULT_CONTENTS.update(cfg)

        # 如果指定了 single-content，则过滤每类 contents，只保留匹配项
        if args.single_content:
            sc = args.single_content
            for k in ["chat_completion", "completion", "embedding"]:
                items = DEFAULT_CONTENTS.get(k, [])
                if sc in items:
                    DEFAULT_CONTENTS[k] = [sc]
                else:
                    # 如果指定的内容不在该类型中则保留原样（不会运行该类型）
                    DEFAULT_CONTENTS[k] = []

        main()
